"""
AnalyzeIt workbook parser service (Week 2 compute layer).

Implements the tool contract the dashboard's controlled tool layer expects:
parse_workbook, profile_dataset, query_dataset, apply_recipe -- plus the
webhook receiver the upload path dispatches to.

Run locally:   venv/bin/uvicorn app.main:app --host 0.0.0.0 --port 8644
"""

from __future__ import annotations

import hashlib
import io
import os
import re
from datetime import datetime
from typing import Any

import duckdb
import openpyxl
import polars as pl
from fastapi import FastAPI, Header, HTTPException, Request
from pydantic import BaseModel

APP_SECRET = os.environ.get("HERMES_WEBHOOK_SECRET", "")
TOOL_SECRET = os.environ.get("TOOL_LAYER_SECRET", "")

app = FastAPI(title="AnalyzeIt Parser Agent", version="0.1.0")

# In-memory dataset store (dataset_id -> parquet bytes + metadata).
# Swap for Supabase Storage / object store in production.
DATASETS: dict[str, dict[str, Any]] = {}


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------

def check_secret(header: str | None, expected: str, kind: str) -> None:
    if not expected:
        raise HTTPException(503, f"{kind} is not configured")
    if header != expected:
        raise HTTPException(401, "Unauthorized")


def money(v: Any) -> float | None:
    """Parse '£1,240.00', '(150.00)', '-410.25', '1,200' -> float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[£$,\s]", "", s.strip("()"))
    if not s or s in {"-", "."}:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def uk_date(v: Any) -> str | None:
    """Normalize dd/mm/yyyy (or datetime cells) to ISO yyyy-mm-dd."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def find_header(rows: list[list[Any]]) -> int:
    """Locate the real header row: a row where >=2 cells are non-null strings
    and at least one looks like 'Date'/'Invoice'/'Name' etc., below any title block."""
    best, best_score = 0, 0
    for i, row in enumerate(rows[:30]):
        cells = [c for c in row if c is not None]
        strings = [c for c in cells if isinstance(c, str)]
        score = len(strings)
        if score >= 2 and i > best_score * 0:  # keep simple scoring
            keywords = sum(
                1 for c in strings
                if str(c).strip().lower() in {
                    "date", "invoice", "supplier", "vendor", "net sales",
                    "vat", "amount", "description", "reference", "total",
                }
            )
            if keywords >= 2 and score > best_score:
                best, best_score = i, score
    return best


def is_junk_row(row: list[Any]) -> bool:
    first = next((c for c in row if c is not None), None)
    if isinstance(first, str):
        f = first.strip().lower()
        if f in {"subtotal", "total", "grand total"} or f.startswith(("*", "this report")):
            return True
    return all(c is None for c in row)


def normalize_vendor(name: Any) -> str:
    s = re.sub(r"\s+", " ", str(name)).strip().lower()
    s = re.sub(r"[.,]$", "", s)
    s = re.sub(r"\b(ltd|limited|ltd\.|co|company)\b", "", s).strip()
    return s


# --------------------------------------------------------------------------
# core parsing
# --------------------------------------------------------------------------

def parse_sheet(xlsx_bytes: bytes) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[0]
    raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header_idx = find_header(raw_rows)
    columns = [str(c).strip() if c is not None else f"col_{j}"
               for j, c in enumerate(raw_rows[header_idx])]
    data_rows = raw_rows[header_idx + 1:]

    records, dropped = [], []
    for r in data_rows:
        if is_junk_row(r):
            dropped.append(r)
            continue
        rec = {}
        for j, col in enumerate(columns):
            v = r[j] if j < len(r) else None
            lc = col.lower()
            if "date" in lc:
                rec[col] = uk_date(v)
            elif any(k in lc for k in ("net", "vat", "amount", "gross", "total", "price", "value")):
                rec[col] = money(v)
            elif any(k in lc for k in ("supplier", "vendor", "name", "customer")):
                rec[col] = normalize_vendor(v) if v else None
            else:
                rec[col] = str(v).strip() if v is not None else None
        # keep only rows with at least one meaningful value beyond an invoice ref
        vals = [x for x in rec.values() if x is not None]
        if len(vals) >= 2:
            records.append(rec)

    df = pl.DataFrame(records) if records else pl.DataFrame({c: [] for c in columns})
    return {
        "dataframe": df,
        "header_row": header_idx,
        "columns": columns,
        "dropped_rows": len(dropped),
        "notes": _extract_notes(wb),
    }


def _extract_notes(wb: openpyxl.Workbook) -> list[dict[str, str]]:
    """Pull Old code / New code mapping pairs from trailing sheets."""
    mappings: list[dict[str, str]] = []
    for ws in wb.worksheets[1:]:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [c for c in row if c is not None]
            if len(cells) == 2 and all(isinstance(c, str) for c in cells):
                a, b = (str(c).strip() for c in cells)
                if i >= 2 and re.match(r"^[A-Z]{2,}[-_]?\w*$", a) and b:
                    mappings.append({"old_code": a, "new_code": b})
    return mappings


def fingerprint(df: pl.DataFrame) -> str:
    h = hashlib.sha256()
    h.update(df.columns.__str__().encode())
    h.update(df.write_csv().encode())
    return h.hexdigest()


# --------------------------------------------------------------------------
# endpoints
# --------------------------------------------------------------------------

@app.get("/health")
def health():
    return {
        "status": "healthy",
        "agent": "AnalyzeIt Parser Agent",
        "datasets": len(DATASETS),
        "time": datetime.utcnow().isoformat(),
    }


class WebhookPayload(BaseModel):
    event: str
    dataset_id: str | None = None
    filename: str | None = None
    tenant_id: str | None = None
    workspace_id: str | None = None
    storage_path: str | None = None
    sha256: str | None = None
    instructions: str | None = None


@app.post("/webhooks/{name}")
async def webhook(name: str, payload: WebhookPayload,
                  x_hermes_secret: str | None = Header(default=None)):
    check_secret(x_hermes_secret, APP_SECRET, "HERMES_WEBHOOK_SECRET")

    result: dict[str, Any] = {"received": payload.event}
    if payload.event == "workbook.uploaded" and payload.storage_path:
        # In production: download from Supabase Storage using storage_path.
        # Locally: the file must be pushed via POST /datasets first.
        ds = DATASETS.get(payload.dataset_id or payload.storage_path)
        if ds:
            parsed = parse_sheet(ds["bytes"])
            df = parsed["dataframe"]
            sig = fingerprint(df)
            DATASETS[payload.dataset_id or payload.storage_path].update({
                "df": df, "source_signature": sig,
                "parsed": {k: v for k, v in parsed.items() if k != "dataframe"},
            })
            result["parse"] = {
                "rows": df.height, "columns": df.columns,
                "header_row": parsed["header_row"],
                "dropped_junk_rows": parsed["dropped_rows"],
                "mappings_found": parsed["notes"],
                "source_signature": sig,
            }
    return result


@app.post("/datasets/{dataset_id}")
async def push_dataset(dataset_id: str, request: Request,
                       x_hermes_secret: str | None = Header(default=None)):
    """Upload raw workbook bytes so webhook/tool calls can parse them."""
    check_secret(x_hermes_secret, APP_SECRET, "HERMES_WEBHOOK_SECRET")
    body = await request.body()
    if not body:
        raise HTTPException(400, "empty body")
    try:
        parse_probe = openpyxl.load_workbook(io.BytesIO(body), read_only=True)
        sheets = parse_probe.sheetnames
    except Exception:
        raise HTTPException(400, "not a valid xlsx workbook")
    DATASETS[dataset_id] = {"bytes": body, "filename": dataset_id, "sheets": sheets}
    return {"stored": True, "dataset_id": dataset_id, "sheets": sheets}


@app.post("/api/v1/tools/{tool}")
async def run_tool(tool: str, request: Request,
                   authorization: str | None = Header(default=None)):
    check_secret(authorization, f"Bearer {TOOL_SECRET}" if TOOL_SECRET else "",
                 "TOOL_LAYER_SECRET")
    body = await request.json()
    params = body.get("params", {})
    dry_run = body.get("dry_run", True)

    if tool == "parse_workbook":
        ds_id = params.get("dataset_id")
        ds = DATASETS.get(ds_id)
        if not ds:
            raise HTTPException(404, f"unknown dataset_id {ds_id}")
        if "df" not in ds:
            parsed = parse_sheet(ds["bytes"])
            df = parsed["dataframe"]
            ds["df"] = df
            ds["source_signature"] = fingerprint(df)
            ds["parsed"] = {k: v for k, v in parsed.items() if k != "dataframe"}
        p = ds["parsed"]
        return {"status": "result", "result": {
            "rows": ds["df"].height, "columns": p["columns"],
            "header_row": p["header_row"],
            "dropped_junk_rows": p["dropped_rows"],
            "mappings": p["notes"],
            "source_signature": ds["source_signature"],
        }, "evidence": {"tool": tool, "dataset_id": ds_id}}

    if tool == "profile_dataset":
        ds = DATASETS.get(params.get("dataset_id"))
        if not ds or "df" not in ds:
            raise HTTPException(404, "dataset not parsed yet; call parse_workbook first")
        df = ds["df"]
        profile = {}
        for col in df.columns:
            s = df[col]
            entry: dict[str, Any] = {
                "dtype": str(s.dtype),
                "nulls": int(s.null_count()),
                "n_unique": int(s.n_unique()),
            }
            if s.dtype.is_numeric():
                entry |= {"sum": float(s.sum()), "mean": float(s.mean()),
                          "min": float(s.min()), "max": float(s.max())}
            profile[col] = entry
        dupes = df.filter(df.is_duplicated())
        return {"status": "result", "result": {
            "rows": df.height, "duplicate_rows": dupes.height, "columns": profile},
            "evidence": {"tool": tool}}

    if tool == "query_dataset":
        ds = DATASETS.get(params.get("dataset_id"))
        if not ds or "df" not in ds:
            raise HTTPException(404, "dataset not parsed yet; call parse_workbook first")
        sql = params.get("sql") or ""
        con = duckdb.connect()
        con.register("ds", ds["df"].to_pandas())
        rows = con.execute(sql).fetchall()
        cols = [d[0] for d in con.execute(sql).description]
        con.close()
        return {"status": "result", "result": {"columns": cols, "rows": [list(r) for r in rows]},
                "evidence": {"tool": tool, "sql": sql}}

    if tool == "apply_recipe":
        ds = DATASETS.get(params.get("dataset_id"))
        if not ds or "df" not in ds:
            raise HTTPException(404, "dataset not parsed yet; call parse_workbook first")
        steps = params.get("steps") or []
        preview_steps = [{"step": s.get("type"), "dry_run": dry_run} for s in steps]
        if not dry_run:
            df = ds["df"]
            for step in steps:
                st = step.get("type")
                if st == "dedupe":
                    df = df.unique(keep="first")
                elif st == "drop_nulls" and step.get("column"):
                    df = df.drop_nulls(subset=[step["column"]])
            ds["cleaned"] = df
        return {"status": "result", "result": {
            "applied_steps": preview_steps if dry_run else [
                {"type": s.get("type"), "done": True} for s in steps],
            "rows_before": ds["df"].height,
            "rows_after": ds["cleaned"].height if "cleaned" in ds else None},
            "execution_metadata": {"dry_run": dry_run}}

    raise HTTPException(404, f"unknown tool '{tool}'")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8644)
