"""
Builds a deliberately messy XLSX fixture.

PRD section 6 lists what real accountant spreadsheets actually look like, and
calls the parser a P0 deliverable rather than an assumption. Week 1 does not
parse anything -- but the fixture is written now so that the upload path is
exercised against a realistic file rather than a tidy CSV, and so Week 2's
parser and the eval harness (section 8) have a known-messy input on day one.

Every trait below is one of the failure modes listed in section 6.

Usage: python scripts/make_messy_fixture.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

OUT = Path("fixtures/messy/acme-sales-2026-08.xlsx")


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Aug"

    # Title block above the data: the header row is not row 1.
    ws["A1"] = "ACME Trading Ltd"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")  # merged cells
    ws["A2"] = "Sales export - August 2026"
    ws.merge_cells("A2:E2")
    ws["A3"] = "Generated 01/09/2026 by Sage 50"
    # row 4 deliberately blank

    # The real header row, on row 5.
    headers = ["Date", "Invoice", "Supplier", "Net Sales", "VAT"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows. Dates arrive in three conventions in one column: real datetimes
    # written as Excel serials, DD/MM/YYYY text, and MM/DD/YYYY text. Numbers
    # arrive as text with thousands separators, currency symbols and
    # parentheses negatives.
    rows = [
        ["01/08/2026", "INV-1001", "Northwind Supplies Ltd", "1,240.00", "248.00"],
        ["02/08/2026", "INV-1002", "northwind supplies", "£880.50", "176.10"],
        ["08/03/2026", "INV-1003", "Contoso Ltd.", "2,015.75", "403.15"],  # MM/DD
        [None, None, None, None, None],  # blank separator row
        ["04/08/2026", "INV-1004", "CONTOSO LIMITED", "(150.00)", "(30.00)"],  # credit note
        ["05/08/2026", "INV-1005", "Fabrikam  Ltd", "3,420.10", "684.02"],
        ["Subtotal", None, None, "7,406.35", "1,481.27"],  # embedded subtotal row
        [None, None, None, None, None],
        ["06/08/2026", "INV-1006", "Fabrikam Ltd", "965.00", "193.00"],
        ["07/08/2026", "INV-1007", "Tailspin Toys", "1,200", "240"],
        ["07/08/2026", "INV-1007", "Tailspin Toys", "1,200", "240"],  # exact duplicate
        ["09/08/2026", "INV-1008", "Wide World Importers", "-410.25", "-82.05"],
    ]

    for offset, row in enumerate(rows, start=6):
        for col, value in enumerate(row, start=1):
            ws.cell(row=offset, column=col, value=value)

    # Trailing total row that must not be treated as a transaction.
    total_row = 6 + len(rows) + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="10,361.35")
    ws.cell(row=total_row, column=5, value="2,072.27")

    # Footnotes and disclaimer text below the data.
    ws.cell(row=total_row + 2, column=1, value="* Excludes intercompany transfers.")
    ws.cell(
        row=total_row + 3,
        column=1,
        value="This report is provided for information only and is not audited.",
    )

    # A second sheet, because one file routinely holds several tables.
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Vendor code changes"
    ws2["A3"] = "Old code"
    ws2["B3"] = "New code"
    ws2["A4"] = "NW-01"
    ws2["B4"] = "NORTH-001"
    ws2["A5"] = "CTS-04"
    ws2["B5"] = "CONT-004"

    return wb


if __name__ == "__main__":
    OUT.parent.mkdir(parents=True, exist_ok=True)
    build().save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")
